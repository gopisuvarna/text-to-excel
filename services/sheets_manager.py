"""
sheets_manager.py — Google Sheets data layer.

Replaces excel_manager.py as the primary storage backend.
Keeps the same public function signatures so app.py requires no changes:

    update_sheets(records)          → { records_added, new_columns_added }
    load_existing_data_from_sheets()→ (list[dict], list[str])
    get_schema_columns()            → list[str]
    export_as_excel()               → bytes  (for /download endpoint)

Authentication: Google Service Account credentials stored as a JSON string
in the environment variable GOOGLE_SERVICE_ACCOUNT_JSON (preferred for
cloud deployments) or as a path in GOOGLE_APPLICATION_CREDENTIALS.

Required env vars:
    GOOGLE_SERVICE_ACCOUNT_JSON   — full JSON string of the service account key
    GOOGLE_SPREADSHEET_ID         — the spreadsheet ID from the URL
    GOOGLE_SHEET_NAME             — worksheet tab name (default: MasterData)
"""

import io
import json
import logging
import os
import time
from typing import Any

import gspread
import pandas as pd
from google.oauth2.service_account import Credentials

from .schema_manager import (
    MASTER_SCHEMA,
    _DISCARD_FIELDS,
    get_new_columns,
    merge_schemas,
    normalize_record,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

_RETRY_ATTEMPTS = 3
_RETRY_BACKOFF  = 2   # seconds, doubles each retry


# ---------------------------------------------------------------------------
# Auth — build gspread client
# ---------------------------------------------------------------------------

def _build_client() -> gspread.Client:
    """
    Authenticate using service account credentials.

    Priority:
    1. GOOGLE_SERVICE_ACCOUNT_JSON env var — full JSON string (required on Render/Vercel)
    2. GOOGLE_APPLICATION_CREDENTIALS env var — path to a local JSON file (local dev only)
    3. credentials.json in the working directory (local dev fallback)
    """
    json_str = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()

    if json_str:
        try:
            info = json.loads(json_str)
        except json.JSONDecodeError as exc:
            raise EnvironmentError(
                "GOOGLE_SERVICE_ACCOUNT_JSON is set but contains invalid JSON. "
                "Make sure you pasted the full contents of credentials.json."
            ) from exc
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        return gspread.authorize(creds)

    # Local dev fallback — file path
    path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "credentials.json").strip()
    if not os.path.exists(path):
        raise EnvironmentError(
            "Google Sheets credentials are not configured.\n"
            "  • On Render: set the GOOGLE_SERVICE_ACCOUNT_JSON environment variable "
            "to the full JSON content of your service account key file.\n"
            "  • Locally: place credentials.json in the project root, or set "
            "GOOGLE_APPLICATION_CREDENTIALS to its path."
        )

    creds = Credentials.from_service_account_file(path, scopes=SCOPES)
    return gspread.authorize(creds)


def _get_worksheet() -> gspread.Worksheet:
    """Open the configured spreadsheet and worksheet, creating the sheet tab if missing."""
    spreadsheet_id = os.getenv("GOOGLE_SPREADSHEET_ID", "").strip()
    if not spreadsheet_id:
        raise EnvironmentError(
            "GOOGLE_SPREADSHEET_ID is not set. Add it to your .env file."
        )

    sheet_name = os.getenv("GOOGLE_SHEET_NAME", "MasterData").strip()
    client      = _build_client()
    spreadsheet = client.open_by_key(spreadsheet_id)

    # Get existing worksheet or create it
    try:
        ws = spreadsheet.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        logger.info("Worksheet '%s' not found — creating it.", sheet_name)
        ws = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=30)
        # Write header row with master schema
        ws.append_row(MASTER_SCHEMA, value_input_option="USER_ENTERED")

    return ws


# ---------------------------------------------------------------------------
# Retry decorator
# ---------------------------------------------------------------------------

def _with_retry(fn, *args, **kwargs):
    """Call fn with simple exponential-backoff retry for transient API errors."""
    last_exc = None
    wait = _RETRY_BACKOFF
    for attempt in range(1, _RETRY_ATTEMPTS + 1):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as exc:
            status = exc.response.status_code if exc.response else 0
            if status == 429 or status >= 500:
                logger.warning("Sheets API error %s (attempt %d/%d), retrying in %ds…",
                               status, attempt, _RETRY_ATTEMPTS, wait)
                time.sleep(wait)
                wait *= 2
                last_exc = exc
            else:
                raise
    raise last_exc


# ---------------------------------------------------------------------------
# Read all existing data
# ---------------------------------------------------------------------------

def load_existing_data_from_sheets() -> tuple[list[dict], list[str]]:
    """
    Fetch all rows from Google Sheets.
    Returns (list_of_row_dicts, list_of_column_names).
    If the sheet is empty/header-only, returns ([], MASTER_SCHEMA).
    """
    ws = _with_retry(_get_worksheet)
    all_values = _with_retry(ws.get_all_values)

    if not all_values:
        logger.info("Sheet is empty — initialising with master schema header.")
        _with_retry(ws.append_row, MASTER_SCHEMA, value_input_option="USER_ENTERED")
        return [], list(MASTER_SCHEMA)

    headers = all_values[0]
    if not headers:
        return [], list(MASTER_SCHEMA)

    # Drop discard columns from the header view
    clean_headers = [h for h in headers if h.lower().strip() not in _DISCARD_FIELDS]

    rows = []
    for raw_row in all_values[1:]:
        # Pad short rows so zip works cleanly
        padded = raw_row + [""] * (len(headers) - len(raw_row))
        row_dict = {headers[i]: (padded[i] if padded[i] != "" else None)
                    for i in range(len(headers))
                    if headers[i].lower().strip() not in _DISCARD_FIELDS}
        rows.append(row_dict)

    logger.info("Loaded %d rows with %d columns from Sheets.", len(rows), len(clean_headers))
    return rows, clean_headers


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def _build_duplicate_key(row: dict) -> str | None:
    tag    = str(row.get("Tag No")    or "").strip()
    serial = str(row.get("Serial No") or "").strip()
    date   = str(row.get("Date")      or "").strip()

    if tag and tag.lower() not in ("none", "nan", "null", ""):
        return f"{tag}|{serial}|{date}"
    if serial and serial.lower() not in ("none", "nan", "null", ""):
        return f"|{serial}|{date}"
    return None


def _existing_keys(rows: list[dict]) -> set[str]:
    keys: set[str] = set()
    for row in rows:
        key = _build_duplicate_key(row)
        if key:
            keys.add(key)
    return keys


# ---------------------------------------------------------------------------
# Column / header management
# ---------------------------------------------------------------------------

def _get_current_headers(ws: gspread.Worksheet) -> list[str]:
    """Read just the first row (headers) from the worksheet."""
    result = _with_retry(ws.row_values, 1)
    return result if result else []


def _ensure_columns(ws: gspread.Worksheet, required_columns: list[str]) -> list[str]:
    """
    Ensure all required_columns exist as headers in row 1.
    Appends any missing columns to the right. Returns the final header list.
    """
    current = _get_current_headers(ws)
    current_set = set(current)
    new_cols = [c for c in required_columns if c not in current_set]

    if new_cols:
        final_headers = current + new_cols
        # Rewrite the full header row
        _with_retry(
            ws.update,
            "A1",
            [final_headers],
            value_input_option="USER_ENTERED",
        )
        logger.info("Added new columns to sheet: %s", new_cols)
        return final_headers

    return current


# ---------------------------------------------------------------------------
# Core write function
# ---------------------------------------------------------------------------

def update_sheets(records: list[dict]) -> dict[str, Any]:
    """
    Normalize records, expand schema if needed, deduplicate, and append
    new rows to Google Sheets.

    Returns: { "records_added": int, "new_columns_added": list[str] }
    """
    if not records:
        return {"records_added": 0, "new_columns_added": []}

    # 1. Normalize all incoming records
    normalized = [normalize_record(r) for r in records]

    # 2. Load existing sheet data
    existing_rows, existing_columns = load_existing_data_from_sheets()

    # 3. Discover new columns across all incoming records
    all_keys: list[str] = []
    for rec in normalized:
        for k in rec.keys():
            if k not in all_keys:
                all_keys.append(k)

    new_cols     = get_new_columns(existing_columns, all_keys)
    final_columns = merge_schemas(existing_columns, normalized)

    # 4. Ensure sheet headers match final_columns (adds missing cols)
    ws = _with_retry(_get_worksheet)
    actual_headers = _ensure_columns(ws, final_columns)

    # 5. Deduplicate
    dup_keys = _existing_keys(existing_rows)
    new_rows: list[dict] = []
    skipped = 0

    for rec in normalized:
        key = _build_duplicate_key(rec)
        if key and key in dup_keys:
            logger.info("Duplicate skipped: %s", key)
            skipped += 1
            continue
        new_rows.append(rec)
        if key:
            dup_keys.add(key)

    logger.info("Records to insert: %d | Duplicates skipped: %d", len(new_rows), skipped)

    if not new_rows:
        return {"records_added": 0, "new_columns_added": new_cols}

    # 6. Convert to row arrays aligned to the final header order
    rows_to_append = []
    for rec in new_rows:
        row = [str(rec.get(col, "") or "") for col in actual_headers]
        rows_to_append.append(row)

    # 7. Batch append — single API call for all rows
    _with_retry(
        ws.append_rows,
        rows_to_append,
        value_input_option="USER_ENTERED",
        insert_data_option="INSERT_ROWS",
    )

    logger.info("Sheets updated — %d rows appended, %d new cols: %s",
                len(new_rows), len(new_cols), new_cols)

    return {"records_added": len(new_rows), "new_columns_added": new_cols}


# ---------------------------------------------------------------------------
# Schema query
# ---------------------------------------------------------------------------

def get_schema_columns() -> list[str]:
    """Return the current column headers from the sheet."""
    ws = _with_retry(_get_worksheet)
    headers = _get_current_headers(ws)
    return headers if headers else list(MASTER_SCHEMA)


# ---------------------------------------------------------------------------
# Stats query
# ---------------------------------------------------------------------------

def get_stats() -> dict[str, Any]:
    """Return aggregate stats from the sheet."""
    rows, columns = load_existing_data_from_sheets()
    new_col_count = len([c for c in columns if c not in MASTER_SCHEMA])
    return {
        "total_records":     len(rows),
        "files_processed":   0,
        "total_columns":     len(columns),
        "new_columns_added": new_col_count,
    }


# ---------------------------------------------------------------------------
# Export as Excel (for /download endpoint)
# ---------------------------------------------------------------------------

def export_as_excel() -> bytes:
    """
    Fetch all data from Google Sheets and return as an in-memory .xlsx bytes object.
    This allows /download to keep returning an Excel file even though the backend
    now uses Sheets as its primary store.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    rows, columns = load_existing_data_from_sheets()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    if not columns:
        columns = list(MASTER_SCHEMA)

    # Header
    ws.append(columns)

    # Data rows
    for row_dict in rows:
        ws.append([str(row_dict.get(col, "") or "") for col in columns])

    num_rows = ws.max_row
    num_cols = ws.max_column

    # Formatting
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    BODY_FONT   = Font(size=10, name="Calibri")
    ALT_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row_idx in range(2, num_rows + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in ws[row_idx]:
            cell.font      = BODY_FONT
            cell.alignment = Alignment(vertical="center")
            if fill.fill_type:
                cell.fill = fill

    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max(
            (len(str(ws[f"{col_letter}{r}"].value or "")) for r in range(1, num_rows + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    if num_rows > 1:
        table_ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
        table = Table(displayName="InventoryTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,  showColumnStripes=False,
        )
        ws._tables.clear()
        ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
