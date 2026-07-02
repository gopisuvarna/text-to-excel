"""
excel_manager.py — Manages loading, updating, formatting, and saving the master Excel workbook.
"""

import logging
import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .schema_manager import (
    MASTER_SCHEMA, _DISCARD_FIELDS,
    get_new_columns, merge_schemas, normalize_record,
)

logger = logging.getLogger(__name__)


def _get_excel_path() -> Path:
    raw = os.getenv("MASTER_EXCEL_PATH", "output/master_inventory.xlsx")
    path = Path(raw)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# Load / create
# ---------------------------------------------------------------------------

def load_existing_data(excel_path: Path) -> tuple[pd.DataFrame, list[str]]:
    """
    Load existing Excel data into a DataFrame.
    Returns (dataframe, columns_list).
    If the file does not exist or is empty, returns an empty DataFrame
    with the MASTER_SCHEMA columns.
    """
    if not excel_path.exists():
        logger.info("No existing Excel file found — will create a new one.")
        empty_df = pd.DataFrame(columns=MASTER_SCHEMA)
        return empty_df, list(MASTER_SCHEMA)

    try:
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl")
        # Replace pandas NaN with None for consistency
        df = df.where(pd.notna(df), other=None)
        # Drop any discard columns that crept in from previous bad uploads
        bad_cols = [c for c in df.columns if c.lower().strip() in _DISCARD_FIELDS]
        if bad_cols:
            logger.warning("Dropping previously persisted discard columns: %s", bad_cols)
            df.drop(columns=bad_cols, inplace=True)
        columns = list(df.columns)
        logger.info("Loaded %d existing records with columns: %s", len(df), columns)
        return df, columns
    except Exception as exc:
        logger.error("Failed to load Excel file '%s': %s — creating fresh workbook.", excel_path, exc)
        empty_df = pd.DataFrame(columns=MASTER_SCHEMA)
        return empty_df, list(MASTER_SCHEMA)


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def _build_duplicate_key(row: pd.Series) -> str | None:
    """
    Build a composite duplicate-check key from Tag No + Serial No + Date.
    Falls back to Serial No + Date, then None (meaning insert always).
    """
    tag = str(row.get("Tag No") or "").strip()
    serial = str(row.get("Serial No") or "").strip()
    date = str(row.get("Date") or "").strip()

    if tag and tag.lower() not in ("none", "nan", "null"):
        return f"{tag}|{serial}|{date}"
    if serial and serial.lower() not in ("none", "nan", "null"):
        return f"|{serial}|{date}"
    return None  # No meaningful key → always insert


def _existing_keys(df: pd.DataFrame) -> set[str]:
    keys: set[str] = set()
    for _, row in df.iterrows():
        key = _build_duplicate_key(row)
        if key:
            keys.add(key)
    return keys


# ---------------------------------------------------------------------------
# Core update function
# ---------------------------------------------------------------------------

def update_excel(records: list[dict]) -> dict[str, Any]:
    """
    Normalize records, expand schema if needed, deduplicate, and append
    new rows to the master Excel workbook.

    Returns a summary dict:
        records_added: int
        new_columns_added: list[str]
    """
    excel_path = _get_excel_path()

    # 1. Normalize records
    normalized = [normalize_record(r) for r in records]

    # 2. Load existing data
    existing_df, existing_columns = load_existing_data(excel_path)

    # 3. Discover new columns
    all_keys: list[str] = []
    for rec in normalized:
        for k in rec.keys():
            if k not in all_keys:
                all_keys.append(k)

    new_cols = get_new_columns(existing_columns, all_keys)
    final_columns = merge_schemas(existing_columns, normalized)

    # 4. Expand existing DataFrame with new columns (fill NaN)
    for col in new_cols:
        existing_df[col] = None

    existing_df = existing_df.reindex(columns=final_columns)

    # 5. Build set of existing duplicate keys
    dup_keys = _existing_keys(existing_df)

    # 6. Filter new records — skip duplicates
    new_rows: list[dict] = []
    skipped = 0
    for rec in normalized:
        row_series = pd.Series(rec)
        key = _build_duplicate_key(row_series)
        if key and key in dup_keys:
            logger.info("Duplicate skipped: %s", key)
            skipped += 1
            continue
        new_rows.append(rec)
        if key:
            dup_keys.add(key)

    logger.info("Records to insert: %d | Duplicates skipped: %d", len(new_rows), skipped)

    if not new_rows:
        logger.info("No new records to insert.")
        return {"records_added": 0, "new_columns_added": new_cols}

    # 7. Build new rows DataFrame, reindexed to final_columns
    new_df = pd.DataFrame(new_rows).reindex(columns=final_columns)

    # 8. Concatenate
    combined_df = pd.concat([existing_df, new_df], ignore_index=True)

    # 9. Drop fully empty rows
    combined_df.dropna(how="all", inplace=True)

    # 10. Save to Excel with formatting
    _save_with_formatting(combined_df, excel_path)

    logger.info(
        "Excel updated — %d records added, %d new columns: %s",
        len(new_rows), len(new_cols), new_cols,
    )

    return {"records_added": len(new_rows), "new_columns_added": new_cols}


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _save_with_formatting(df: pd.DataFrame, path: Path) -> None:
    """Write DataFrame to Excel and apply professional table formatting."""
    # Write raw data first via pandas (fast)
    df.to_excel(str(path), index=False, engine="openpyxl")

    # Re-open with openpyxl to apply formatting
    wb = load_workbook(str(path))
    ws = wb.active
    ws.title = "Inventory"

    num_rows = ws.max_row
    num_cols = ws.max_column

    # ── Header row ──────────────────────────────────────────────────────────
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # ── Body rows ────────────────────────────────────────────────────────────
    for row_idx in range(2, num_rows + 1):
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()  # alternate shading
        for cell in ws[row_idx]:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="center")
            if fill.fill_type:
                cell.fill = fill

    # ── Auto-adjust column widths ────────────────────────────────────────────
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, num_rows + 1):
            cell_val = ws[f"{col_letter}{row_idx}"].value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        adjusted = min(max_len + 4, 50)  # cap at 50
        ws.column_dimensions[col_letter].width = adjusted

    # ── Freeze header row ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Excel Table (structured reference, auto-filter) ──────────────────────
    if num_rows > 1:
        table_ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
        table = Table(displayName="InventoryTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # Remove existing tables to avoid duplicates on re-save
        ws._tables.clear()
        ws.add_table(table)

    wb.save(str(path))
    logger.info("Excel saved with formatting: %s", path)
